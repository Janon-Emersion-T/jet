from pathlib import Path
import csv
from datetime import datetime
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path

SUPPORTED_EXTENSIONS = [".csv", ".xlsx"]
MAX_FILES = 25
MAX_ROWS_TO_SCAN = 1500

TASK_KEYWORDS = [
    "task",
    "tasks",
    "todo",
    "employee",
    "staff",
    "work",
    "assignment",
]

EMPLOYEE_KEYWORDS = ["employee", "staff", "assignee", "assigned", "owner", "name"]
TASK_NAME_KEYWORDS = ["task", "title", "work", "assignment", "todo", "description"]
STATUS_KEYWORDS = ["status", "progress", "state"]
DUE_DATE_KEYWORDS = ["due", "deadline", "target date", "end date"]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/reports",
        "storage/presentations",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_task_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            name = file.name.lower()
            if any(keyword in name for keyword in TASK_KEYWORDS):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _normalize(value):
    return str(value or "").strip()


def _normalize_lower(value):
    return _normalize(value).lower()


def _find_columns(headers, keywords):
    indexes = []

    for index, header in enumerate(headers):
        header_text = _normalize_lower(header)
        if any(keyword in header_text for keyword in keywords):
            indexes.append(index)

    return indexes


def _parse_date(value):
    if value is None:
        return None

    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            return None

    text = _normalize(value)
    if not text:
        return None

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _is_done(status):
    status = _normalize_lower(status)
    return any(word in status for word in ["done", "completed", "complete", "closed", "finished"])


def _is_in_progress(status):
    status = _normalize_lower(status)
    return any(word in status for word in ["progress", "doing", "working", "active", "started"])


def _is_blocked(status):
    status = _normalize_lower(status)
    return any(word in status for word in ["blocked", "hold", "stuck", "waiting"])


def _analyze_task_rows(headers, rows):
    employee_columns = _find_columns(headers, EMPLOYEE_KEYWORDS)
    task_columns = _find_columns(headers, TASK_NAME_KEYWORDS)
    status_columns = _find_columns(headers, STATUS_KEYWORDS)
    due_columns = _find_columns(headers, DUE_DATE_KEYWORDS)

    total_tasks = 0
    completed = 0
    in_progress = 0
    blocked = 0
    pending = 0
    overdue = 0
    missing_assignee = 0
    missing_task_name = 0
    missing_due_date = 0

    today = datetime.now().date()
    assignee_counts = {}
    warnings = []

    employee_index = employee_columns[0] if employee_columns else None
    task_index = task_columns[0] if task_columns else None
    status_index = status_columns[0] if status_columns else None
    due_index = due_columns[0] if due_columns else None

    for row_number, row in enumerate(rows, start=2):
        if not any(_normalize(cell) for cell in row):
            continue

        total_tasks += 1

        assignee = _normalize(row[employee_index]) if employee_index is not None and employee_index < len(row) else ""
        task_name = _normalize(row[task_index]) if task_index is not None and task_index < len(row) else ""
        status = _normalize(row[status_index]) if status_index is not None and status_index < len(row) else ""
        due_date = _parse_date(row[due_index]) if due_index is not None and due_index < len(row) else None

        if assignee:
            assignee_counts[assignee] = assignee_counts.get(assignee, 0) + 1
        else:
            missing_assignee += 1
            warnings.append(f"- Row {row_number}: Missing employee/assignee.")

        if not task_name:
            missing_task_name += 1
            warnings.append(f"- Row {row_number}: Missing task title/description.")

        if due_index is not None and due_date is None:
            missing_due_date += 1
            warnings.append(f"- Row {row_number}: Missing or unreadable due date.")

        if _is_done(status):
            completed += 1
        elif _is_in_progress(status):
            in_progress += 1
        elif _is_blocked(status):
            blocked += 1
        else:
            pending += 1

        if due_date and due_date < today and not _is_done(status):
            overdue += 1
            warnings.append(f"- Row {row_number}: Overdue task detected.")

    return {
        "employee_columns": [headers[i] for i in employee_columns if i < len(headers)],
        "task_columns": [headers[i] for i in task_columns if i < len(headers)],
        "status_columns": [headers[i] for i in status_columns if i < len(headers)],
        "due_columns": [headers[i] for i in due_columns if i < len(headers)],
        "total_tasks": total_tasks,
        "completed": completed,
        "in_progress": in_progress,
        "blocked": blocked,
        "pending": pending,
        "overdue": overdue,
        "missing_assignee": missing_assignee,
        "missing_task_name": missing_task_name,
        "missing_due_date": missing_due_date,
        "assignee_counts": assignee_counts,
        "warnings": warnings[:60],
    }


def _format_summary(file_type, summary):
    lines = [
        f"- Type: {file_type}",
        f"- Total tasks detected: {summary['total_tasks']}",
        f"- Completed: {summary['completed']}",
        f"- In progress: {summary['in_progress']}",
        f"- Pending/unclear: {summary['pending']}",
        f"- Blocked: {summary['blocked']}",
        f"- Overdue: {summary['overdue']}",
        f"- Missing assignee rows: {summary['missing_assignee']}",
        f"- Missing task title rows: {summary['missing_task_name']}",
        f"- Missing due date rows: {summary['missing_due_date']}",
        f"- Employee columns: {', '.join(map(str, summary['employee_columns'])) if summary['employee_columns'] else 'NONE'}",
        f"- Task columns: {', '.join(map(str, summary['task_columns'])) if summary['task_columns'] else 'NONE'}",
        f"- Status columns: {', '.join(map(str, summary['status_columns'])) if summary['status_columns'] else 'NONE'}",
        f"- Due date columns: {', '.join(map(str, summary['due_columns'])) if summary['due_columns'] else 'NONE'}",
    ]

    if summary["assignee_counts"]:
        lines.append("")
        lines.append("Tasks by assignee:")
        for name, count in sorted(summary["assignee_counts"].items(), key=lambda item: item[0].lower())[:30]:
            lines.append(f"- {name}: {count}")

    if summary["warnings"]:
        lines.append("")
        lines.append("Task tracker warnings:")
        lines.extend(summary["warnings"])
    else:
        lines.append("- No obvious task tracker issues detected.")

    return lines


def _analyze_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]
    summary = _analyze_task_rows(headers, data_rows)

    return _format_summary("CSV", summary)


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        "- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        summary = _analyze_task_rows(headers, data_rows)

        for item in _format_summary("XLSX Sheet", summary):
            lines.append(f"  {item}")

    workbook.close()
    return lines


def employee_task_tracker():
    project, error = _project()
    if error:
        return error

    files = _find_task_files(project)

    lines = [
        "EMPLOYEE TASK TRACKER — PHASE 348",
        f"Project: {project}",
        "",
        "Mode: read-only task tracker inspection.",
        "",
    ]

    if not files:
        lines.append("No employee task spreadsheet files found.")
        lines.append("")
        lines.append("File names should include words like:")
        lines.append("- task")
        lines.append("- tasks")
        lines.append("- todo")
        lines.append("- employee")
        lines.append("- staff")
        return "\n".join(lines)

    lines.append(f"Employee task files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Important:")
    lines.append("- This is a read-only employee task tracker.")
    lines.append("- It does not assign, edit, close, or delete tasks.")
    lines.append("- Use it for inspection before building confirm-based task actions.")

    return "\n".join(lines)
