from pathlib import Path
import csv
from statistics import mean, pstdev
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path


SUPPORTED_EXTENSIONS = [".xlsx", ".csv"]
MAX_FILES = 30
MAX_ROWS_TO_SCAN = 1500

FINANCIAL_FILENAME_KEYWORDS = [
    "finance",
    "financial",
    "sales",
    "expense",
    "income",
    "invoice",
    "payment",
    "ledger",
    "report",
    "transaction",
    "account",
]

AMOUNT_KEYWORDS = [
    "amount",
    "total",
    "value",
    "price",
    "cost",
    "debit",
    "credit",
    "balance",
]

DATE_KEYWORDS = [
    "date",
    "created",
    "paid",
    "invoice date",
    "transaction date",
]

TYPE_KEYWORDS = [
    "type",
    "category",
    "transaction type",
    "account type",
]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/presentations",
        "storage/reports",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_financial_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            lower_name = file.name.lower()
            if any(word in lower_name for word in FINANCIAL_FILENAME_KEYWORDS):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _normalize(value):
    return str(value or "").strip()


def _normalize_lower(value):
    return _normalize(value).lower()


def _to_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = (
        text.replace(",", "")
        .replace("Rs.", "")
        .replace("LKR", "")
        .replace("$", "")
        .strip()
    )

    try:
        return float(text)
    except ValueError:
        return None


def _find_columns(headers, keywords):
    indexes = []

    for index, header in enumerate(headers):
        header_text = _normalize_lower(header)
        if any(keyword in header_text for keyword in keywords):
            indexes.append(index)

    return indexes


def _row_signature(row):
    return "|".join(_normalize(cell).lower() for cell in row)


def _detect_anomalies(headers, rows):
    amount_columns = _find_columns(headers, AMOUNT_KEYWORDS)
    date_columns = _find_columns(headers, DATE_KEYWORDS)
    type_columns = _find_columns(headers, TYPE_KEYWORDS)

    anomalies = []
    numeric_amounts = []
    row_signatures = {}

    for row_number, row in enumerate(rows, start=2):
        signature = _row_signature(row)

        if signature in row_signatures:
            anomalies.append(
                f"- Row {row_number}: Possible duplicate transaction. Matches row {row_signatures[signature]}."
            )
        else:
            row_signatures[signature] = row_number

        if not amount_columns:
            continue

        row_amounts = []

        for column_index in amount_columns:
            value = row[column_index] if column_index < len(row) else None
            number = _to_number(value)

            if number is None:
                anomalies.append(
                    f"- Row {row_number}: Missing or invalid amount in column `{headers[column_index]}`."
                )
                continue

            row_amounts.append(number)
            numeric_amounts.append(abs(number))

            if number == 0:
                anomalies.append(
                    f"- Row {row_number}: Zero-value transaction detected in `{headers[column_index]}`."
                )

            if number < 0:
                anomalies.append(
                    f"- Row {row_number}: Negative amount detected in `{headers[column_index]}`."
                )

        if date_columns:
            for column_index in date_columns:
                value = row[column_index] if column_index < len(row) else None
                if _normalize(value) == "":
                    anomalies.append(
                        f"- Row {row_number}: Missing date in column `{headers[column_index]}`."
                    )

        if type_columns:
            for column_index in type_columns:
                value = _normalize_lower(row[column_index] if column_index < len(row) else "")
                if value == "":
                    anomalies.append(
                        f"- Row {row_number}: Missing transaction type/category in `{headers[column_index]}`."
                    )

        if row_amounts and type_columns:
            row_type_text = " ".join(
                _normalize_lower(row[index] if index < len(row) else "")
                for index in type_columns
            )

            has_income_word = any(word in row_type_text for word in ["income", "revenue", "sale", "credit"])
            has_expense_word = any(word in row_type_text for word in ["expense", "debit", "payment", "purchase", "cost"])

            if not has_income_word and not has_expense_word:
                anomalies.append(
                    f"- Row {row_number}: Unclear transaction type/category `{row_type_text}`."
                )

    if numeric_amounts:
        avg = mean(numeric_amounts)
        deviation = pstdev(numeric_amounts) if len(numeric_amounts) > 1 else 0

        if deviation > 0:
            threshold = avg + (deviation * 2.5)

            for row_number, row in enumerate(rows, start=2):
                for column_index in amount_columns:
                    value = row[column_index] if column_index < len(row) else None
                    number = _to_number(value)

                    if number is not None and abs(number) > threshold:
                        anomalies.append(
                            f"- Row {row_number}: Unusually large amount `{number:,.2f}` compared with scanned transactions."
                        )

    summary = {
        "rows_scanned": len(rows),
        "amount_columns": [headers[index] for index in amount_columns if index < len(headers)],
        "date_columns": [headers[index] for index in date_columns if index < len(headers)],
        "type_columns": [headers[index] for index in type_columns if index < len(headers)],
        "anomaly_count": len(anomalies),
        "anomalies": anomalies[:80],
    }

    return summary


def _format_summary(file_type, headers, summary):
    lines = [
        f"- Type: {file_type}",
        f"- Rows scanned: {summary['rows_scanned']}",
        f"- Amount columns: {', '.join(map(str, summary['amount_columns'])) if summary['amount_columns'] else 'NONE'}",
        f"- Date columns: {', '.join(map(str, summary['date_columns'])) if summary['date_columns'] else 'NONE'}",
        f"- Type/category columns: {', '.join(map(str, summary['type_columns'])) if summary['type_columns'] else 'NONE'}",
        f"- Anomalies found: {summary['anomaly_count']}",
    ]

    if not summary["amount_columns"]:
        lines.append("- Critical warning: No amount-like column detected.")

    if summary["anomalies"]:
        lines.append("")
        lines.append("Detected anomalies:")
        lines.extend(summary["anomalies"])
    else:
        lines.append("- No obvious accounting anomalies detected.")

    return lines


def _analyze_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]
    summary = _detect_anomalies(headers, data_rows)

    return _format_summary("CSV", headers, summary)


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        "- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        summary = _detect_anomalies(headers, data_rows)

        for item in _format_summary("XLSX Sheet", headers, summary):
            lines.append(f"  {item}")

    workbook.close()
    return lines


def accounting_anomaly_detector():
    project, error = _project()
    if error:
        return error

    files = _find_financial_files(project)

    lines = [
        "ACCOUNTING ANOMALY DETECTOR — PHASE 342",
        f"Project: {project}",
        "",
        "Mode: read-only anomaly inspection.",
        "",
    ]

    if not files:
        lines.append("No financial spreadsheet files found.")
        lines.append("")
        lines.append("File names should include words like:")
        lines.append("- sales")
        lines.append("- expense")
        lines.append("- income")
        lines.append("- invoice")
        lines.append("- payment")
        lines.append("- ledger")
        lines.append("- transaction")
        return "\n".join(lines)

    lines.append(f"Financial files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Important:")
    lines.append("- This is a risk-screening assistant, not a certified audit.")
    lines.append("- It flags suspicious patterns for human review.")
    lines.append("- No files were modified.")

    return "\n".join(lines)
