from pathlib import Path
import csv
from collections import Counter

from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path


SUPPORTED_EXTENSIONS = [".xlsx", ".csv"]
MAX_FILES = 20
MAX_ROWS_TO_SCAN = 500


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/presentations",
        "storage/reports",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_spreadsheets(project: Path):
    files = []
    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if not _skip(file) and file.is_file():
                files.append(file)
    return sorted(files)[:MAX_FILES]


def _analyze_csv(file: Path):
    rows = []
    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]

    lines = [
        f"- Type: CSV",
        f"- Columns: {len(headers)}",
        f"- Rows scanned: {len(data_rows)}",
        f"- Headers: {', '.join(headers[:20]) if headers else 'No headers detected'}",
    ]

    missing = 0
    for row in data_rows:
        missing += sum(1 for cell in row if str(cell).strip() == "")

    lines.append(f"- Empty cells found: {missing}")

    if data_rows:
        duplicate_count = len(data_rows) - len({tuple(row) for row in data_rows})
        lines.append(f"- Duplicate rows found: {duplicate_count}")

    return lines


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        f"- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
        f"- Sheet names: {', '.join(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")
        lines.append(f"  - Max rows reported: {sheet.max_row}")
        lines.append(f"  - Max columns reported: {sheet.max_column}")
        lines.append(f"  - Rows scanned: {len(rows)}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        lines.append(f"  - Headers: {', '.join(headers[:20]) if headers else 'No headers detected'}")

        missing = 0
        numeric_values = []

        for row in rows[1:]:
            for cell in row:
                if cell is None or str(cell).strip() == "":
                    missing += 1
                elif isinstance(cell, (int, float)):
                    numeric_values.append(cell)

        lines.append(f"  - Empty cells found: {missing}")

        if numeric_values:
            lines.append(f"  - Numeric min: {min(numeric_values)}")
            lines.append(f"  - Numeric max: {max(numeric_values)}")

        normalized_rows = [tuple(str(cell).strip() for cell in row) for row in rows[1:]]
        duplicate_count = len(normalized_rows) - len(set(normalized_rows))
        lines.append(f"  - Duplicate rows found: {duplicate_count}")

    workbook.close()
    return lines


def spreadsheet_analysis_engine():
    project, error = _project()
    if error:
        return error

    files = _find_spreadsheets(project)

    lines = [
        "SPREADSHEET ANALYSIS ENGINE — PHASE 340",
        f"Project: {project}",
        "",
        "Mode: read-only inspection.",
        "",
    ]

    if not files:
        lines.append("No spreadsheet files found in the current project.")
        lines.append("")
        lines.append("Supported formats:")
        lines.append("- .xlsx")
        lines.append("- .csv")
        return "\n".join(lines)

    lines.append(f"Spreadsheet files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Safety:")
    lines.append("- No spreadsheet files were modified.")
    lines.append("- No formulas were executed manually.")
    lines.append("- Only metadata and cell values were inspected.")

    return "\n".join(lines)
