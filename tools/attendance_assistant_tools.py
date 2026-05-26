from pathlib import Path
import csv
from datetime import datetime, time
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path

SUPPORTED_EXTENSIONS = [".csv", ".xlsx"]
MAX_FILES = 25
MAX_ROWS_TO_SCAN = 2000

ATTENDANCE_KEYWORDS = [
    "attendance",
    "attend",
    "checkin",
    "check-in",
    "checkout",
    "check-out",
    "timesheet",
    "time",
    "staff",
    "employee",
]

EMPLOYEE_KEYWORDS = ["employee", "staff", "name", "worker", "person"]
DATE_KEYWORDS = ["date", "day"]
CHECK_IN_KEYWORDS = ["check in", "check-in", "checkin", "in time", "time in", "start"]
CHECK_OUT_KEYWORDS = ["check out", "check-out", "checkout", "out time", "time out", "end"]
STATUS_KEYWORDS = ["status", "attendance", "present", "absent"]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/reports",
        "storage/presentations",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_attendance_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            name = file.name.lower()
            if any(keyword in name for keyword in ATTENDANCE_KEYWORDS):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _normalize(value):
    return str(value or "").strip()


def _normalize_lower(value):
    return _normalize(value).lower()


def _find_columns(headers, keywords):
    indexes = []

    for index, header in enumerate(headers):
        header_text = _normalize_lower(header)
        if any(keyword in header_text for keyword in keywords):
            indexes.append(index)

    return indexes


def _parse_time(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.time()

    if isinstance(value, time):
        return value

    text = _normalize(value)
    if not text:
        return None

    formats = [
        "%H:%M",
        "%H:%M:%S",
        "%I:%M %p",
        "%I:%M:%S %p",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue

    return None


def _hours_between(start, end):
    if not start or not end:
        return None

    today = datetime.now().date()
    start_dt = datetime.combine(today, start)
    end_dt = datetime.combine(today, end)

    if end_dt < start_dt:
        return None

    seconds = (end_dt - start_dt).total_seconds()
    return seconds / 3600


def _is_absent(status):
    status = _normalize_lower(status)
    return any(word in status for word in ["absent", "leave", "no show", "not present"])


def _is_present(status):
    status = _normalize_lower(status)
    return any(word in status for word in ["present", "attended", "worked"])


def _analyze_attendance_rows(headers, rows):
    employee_columns = _find_columns(headers, EMPLOYEE_KEYWORDS)
    date_columns = _find_columns(headers, DATE_KEYWORDS)
    check_in_columns = _find_columns(headers, CHECK_IN_KEYWORDS)
    check_out_columns = _find_columns(headers, CHECK_OUT_KEYWORDS)
    status_columns = _find_columns(headers, STATUS_KEYWORDS)

    employee_index = employee_columns[0] if employee_columns else None
    check_in_index = check_in_columns[0] if check_in_columns else None
    check_out_index = check_out_columns[0] if check_out_columns else None
    status_index = status_columns[0] if status_columns else None

    total_rows = 0
    present = 0
    absent = 0
    missing_check_in = 0
    missing_check_out = 0
    late_count = 0
    short_shift_count = 0
    total_hours = 0.0
    hour_rows = 0
    employee_counts = {}
    warnings = []

    late_after = time(9, 15)
    minimum_shift_hours = 4

    for row_number, row in enumerate(rows, start=2):
        if not any(_normalize(cell) for cell in row):
            continue

        total_rows += 1

        employee = _normalize(row[employee_index]) if employee_index is not None and employee_index < len(row) else ""
        check_in = _parse_time(row[check_in_index]) if check_in_index is not None and check_in_index < len(row) else None
        check_out = _parse_time(row[check_out_index]) if check_out_index is not None and check_out_index < len(row) else None
        status = _normalize(row[status_index]) if status_index is not None and status_index < len(row) else ""

        if employee:
            employee_counts[employee] = employee_counts.get(employee, 0) + 1

        if _is_absent(status):
            absent += 1
        elif _is_present(status) or check_in:
            present += 1

        if not check_in and not _is_absent(status):
            missing_check_in += 1
            warnings.append(f"- Row {row_number}: Missing check-in time.")

        if not check_out and not _is_absent(status):
            missing_check_out += 1
            warnings.append(f"- Row {row_number}: Missing check-out time.")

        if check_in and check_in > late_after:
            late_count += 1
            warnings.append(f"- Row {row_number}: Late check-in detected.")

        hours = _hours_between(check_in, check_out)
        if hours is not None:
            total_hours += hours
            hour_rows += 1

            if hours < minimum_shift_hours and not _is_absent(status):
                short_shift_count += 1
                warnings.append(f"- Row {row_number}: Short shift detected ({hours:.2f} hours).")

    average_hours = total_hours / hour_rows if hour_rows else 0.0

    return {
        "employee_columns": [headers[i] for i in employee_columns if i < len(headers)],
        "date_columns": [headers[i] for i in date_columns if i < len(headers)],
        "check_in_columns": [headers[i] for i in check_in_columns if i < len(headers)],
        "check_out_columns": [headers[i] for i in check_out_columns if i < len(headers)],
        "status_columns": [headers[i] for i in status_columns if i < len(headers)],
        "total_rows": total_rows,
        "present": present,
        "absent": absent,
        "missing_check_in": missing_check_in,
        "missing_check_out": missing_check_out,
        "late_count": late_count,
        "short_shift_count": short_shift_count,
        "total_hours": total_hours,
        "average_hours": average_hours,
        "employee_counts": employee_counts,
        "warnings": warnings[:80],
    }


def _format_summary(file_type, summary):
    lines = [
        f"- Type: {file_type}",
        f"- Attendance rows detected: {summary['total_rows']}",
        f"- Present/checked-in rows: {summary['present']}",
        f"- Absent/leave rows: {summary['absent']}",
        f"- Missing check-in rows: {summary['missing_check_in']}",
        f"- Missing check-out rows: {summary['missing_check_out']}",
        f"- Late check-ins: {summary['late_count']}",
        f"- Short shifts: {summary['short_shift_count']}",
        f"- Estimated total hours: {summary['total_hours']:,.2f}",
        f"- Average hours per valid shift: {summary['average_hours']:,.2f}",
        f"- Employee columns: {', '.join(map(str, summary['employee_columns'])) if summary['employee_columns'] else 'NONE'}",
        f"- Date columns: {', '.join(map(str, summary['date_columns'])) if summary['date_columns'] else 'NONE'}",
        f"- Check-in columns: {', '.join(map(str, summary['check_in_columns'])) if summary['check_in_columns'] else 'NONE'}",
        f"- Check-out columns: {', '.join(map(str, summary['check_out_columns'])) if summary['check_out_columns'] else 'NONE'}",
        f"- Status columns: {', '.join(map(str, summary['status_columns'])) if summary['status_columns'] else 'NONE'}",
    ]

    if summary["employee_counts"]:
        lines.append("")
        lines.append("Attendance rows by employee:")
        for name, count in sorted(summary["employee_counts"].items(), key=lambda item: item[0].lower())[:30]:
            lines.append(f"- {name}: {count}")

    if summary["warnings"]:
        lines.append("")
        lines.append("Attendance warnings:")
        lines.extend(summary["warnings"])
    else:
        lines.append("- No obvious attendance issues detected.")

    return lines


def _analyze_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]
    summary = _analyze_attendance_rows(headers, data_rows)

    return _format_summary("CSV", summary)


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        "- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        summary = _analyze_attendance_rows(headers, data_rows)

        for item in _format_summary("XLSX Sheet", summary):
            lines.append(f"  {item}")

    workbook.close()
    return lines


def attendance_assistant():
    project, error = _project()
    if error:
        return error

    files = _find_attendance_files(project)

    lines = [
        "ATTENDANCE ASSISTANT — PHASE 349",
        f"Project: {project}",
        "",
        "Mode: read-only attendance inspection.",
        "",
    ]

    if not files:
        lines.append("No attendance spreadsheet files found.")
        lines.append("")
        lines.append("File names should include words like:")
        lines.append("- attendance")
        lines.append("- checkin")
        lines.append("- checkout")
        lines.append("- timesheet")
        lines.append("- employee")
        lines.append("- staff")
        return "\n".join(lines)

    lines.append(f"Attendance files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Important:")
    lines.append("- This is a read-only attendance assistant.")
    lines.append("- It does not modify attendance records.")
    lines.append("- Late and short-shift rules are assistant-level assumptions.")

    return "\n".join(lines)
