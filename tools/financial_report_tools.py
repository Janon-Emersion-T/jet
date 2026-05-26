from pathlib import Path
import csv
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path


SUPPORTED_EXTENSIONS = [".xlsx", ".csv"]
MAX_FILES = 30
MAX_ROWS_TO_SCAN = 1000

INCOME_KEYWORDS = [
    "income",
    "revenue",
    "sales",
    "credit",
    "received",
    "deposit",
]

EXPENSE_KEYWORDS = [
    "expense",
    "cost",
    "debit",
    "payment",
    "paid",
    "purchase",
    "salary",
    "rent",
    "bill",
]

AMOUNT_KEYWORDS = [
    "amount",
    "total",
    "value",
    "price",
    "cost",
    "debit",
    "credit",
    "balance",
]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/presentations",
        "storage/reports",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_financial_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            lower_name = file.name.lower()
            if any(word in lower_name for word in [
                "finance",
                "financial",
                "sales",
                "expense",
                "income",
                "invoice",
                "payment",
                "ledger",
                "report",
                "transaction",
                "account",
            ]):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _to_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = (
        text.replace(",", "")
        .replace("Rs.", "")
        .replace("LKR", "")
        .replace("$", "")
        .strip()
    )

    try:
        return float(text)
    except ValueError:
        return None


def _normalize_header(value):
    return str(value or "").strip().lower()


def _find_amount_columns(headers):
    amount_indexes = []

    for index, header in enumerate(headers):
        header_text = _normalize_header(header)
        if any(keyword in header_text for keyword in AMOUNT_KEYWORDS):
            amount_indexes.append(index)

    return amount_indexes


def _classify_row(headers, row):
    combined = " ".join(str(cell or "").lower() for cell in list(headers) + list(row))

    income_score = sum(1 for word in INCOME_KEYWORDS if word in combined)
    expense_score = sum(1 for word in EXPENSE_KEYWORDS if word in combined)

    if income_score > expense_score:
        return "income"

    if expense_score > income_score:
        return "expense"

    return "unknown"


def _analyze_rows(headers, rows):
    amount_columns = _find_amount_columns(headers)

    if not amount_columns:
        return {
            "rows": len(rows),
            "amount_columns": [],
            "income": 0.0,
            "expense": 0.0,
            "unknown": 0.0,
            "transactions": 0,
            "missing_amounts": 0,
        }

    income_total = 0.0
    expense_total = 0.0
    unknown_total = 0.0
    transactions = 0
    missing_amounts = 0

    for row in rows:
        row_type = _classify_row(headers, row)

        row_amounts = []
        for column_index in amount_columns:
            if column_index < len(row):
                number = _to_number(row[column_index])
                if number is not None:
                    row_amounts.append(abs(number))

        if not row_amounts:
            missing_amounts += 1
            continue

        amount = sum(row_amounts)
        transactions += 1

        if row_type == "income":
            income_total += amount
        elif row_type == "expense":
            expense_total += amount
        else:
            unknown_total += amount

    return {
        "rows": len(rows),
        "amount_columns": amount_columns,
        "income": income_total,
        "expense": expense_total,
        "unknown": unknown_total,
        "transactions": transactions,
        "missing_amounts": missing_amounts,
    }


def _analyze_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]
    result = _analyze_rows(headers, data_rows)

    return _format_result("CSV", headers, result)


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        "- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        result_lines = _format_result("XLSX Sheet", headers, _analyze_rows(headers, data_rows))

        for item in result_lines:
            lines.append(f"  {item}")

    workbook.close()
    return lines


def _format_money(value):
    return f"{value:,.2f}"


def _format_result(file_type, headers, result):
    amount_columns = result["amount_columns"]
    amount_names = []

    for index in amount_columns:
        if index < len(headers):
            amount_names.append(str(headers[index]))

    net = result["income"] - result["expense"]

    lines = [
        f"- Type: {file_type}",
        f"- Rows scanned: {result['rows']}",
        f"- Transactions with amounts: {result['transactions']}",
        f"- Missing/unreadable amount rows: {result['missing_amounts']}",
    ]

    if amount_names:
        lines.append(f"- Amount columns detected: {', '.join(amount_names)}")
    else:
        lines.append("- Amount columns detected: NONE")

    lines.extend([
        f"- Estimated income total: {_format_money(result['income'])}",
        f"- Estimated expense total: {_format_money(result['expense'])}",
        f"- Uncategorized amount total: {_format_money(result['unknown'])}",
        f"- Estimated net position: {_format_money(net)}",
    ])

    if not amount_names:
        lines.append("- Recommendation: Add a clear Amount, Debit, Credit, Total, or Balance column.")

    if result["unknown"] > 0:
        lines.append("- Recommendation: Add a transaction type/category column such as income or expense.")

    return lines


def financial_report_assistant():
    project, error = _project()
    if error:
        return error

    files = _find_financial_files(project)

    lines = [
        "FINANCIAL REPORT ASSISTANT — PHASE 341",
        f"Project: {project}",
        "",
        "Mode: read-only financial inspection.",
        "",
    ]

    if not files:
        lines.append("No financial spreadsheet files found.")
        lines.append("")
        lines.append("File names should include words like:")
        lines.append("- sales")
        lines.append("- expense")
        lines.append("- income")
        lines.append("- invoice")
        lines.append("- payment")
        lines.append("- ledger")
        lines.append("- transaction")
        return "\n".join(lines)

    lines.append(f"Financial files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Important:")
    lines.append("- This is an assistant-level estimate, not a certified accounting report.")
    lines.append("- It depends on column names and available spreadsheet structure.")
    lines.append("- No files were modified.")

    return "\n".join(lines)
