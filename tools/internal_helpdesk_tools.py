from pathlib import Path
import csv
from datetime import datetime
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path

SUPPORTED_EXTENSIONS = [".csv", ".xlsx"]
MAX_FILES = 25
MAX_ROWS_TO_SCAN = 2000

HELPDESK_KEYWORDS = [
    "helpdesk",
    "support",
    "ticket",
    "tickets",
    "issue",
    "issues",
    "request",
    "requests",
]

TICKET_ID_KEYWORDS = ["ticket", "id", "reference", "case"]
SUBJECT_KEYWORDS = ["subject", "title", "issue", "problem", "request", "description"]
REQUESTER_KEYWORDS = ["requester", "employee", "staff", "user", "name", "customer"]
STATUS_KEYWORDS = ["status", "state", "progress"]
PRIORITY_KEYWORDS = ["priority", "severity", "impact", "level"]
DATE_KEYWORDS = ["date", "created", "opened", "reported"]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/reports",
        "storage/presentations",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_helpdesk_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            name = file.name.lower()
            if any(keyword in name for keyword in HELPDESK_KEYWORDS):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _normalize(value):
    return str(value or "").strip()


def _normalize_lower(value):
    return _normalize(value).lower()


def _find_columns(headers, keywords):
    indexes = []

    for index, header in enumerate(headers):
        header_text = _normalize_lower(header)
        if any(keyword in header_text for keyword in keywords):
            indexes.append(index)

    return indexes


def _parse_date(value):
    if value is None:
        return None

    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            return None

    text = _normalize(value)
    if not text:
        return None

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _is_closed(status):
    status = _normalize_lower(status)
    return any(word in status for word in ["closed", "resolved", "done", "completed", "fixed"])


def _is_open(status):
    status = _normalize_lower(status)
    return any(word in status for word in ["open", "new", "pending", "waiting", "in progress", "active"])


def _is_high_priority(priority):
    priority = _normalize_lower(priority)
    return any(word in priority for word in ["high", "urgent", "critical", "severe", "p1", "p0"])


def _analyze_helpdesk_rows(headers, rows):
    ticket_columns = _find_columns(headers, TICKET_ID_KEYWORDS)
    subject_columns = _find_columns(headers, SUBJECT_KEYWORDS)
    requester_columns = _find_columns(headers, REQUESTER_KEYWORDS)
    status_columns = _find_columns(headers, STATUS_KEYWORDS)
    priority_columns = _find_columns(headers, PRIORITY_KEYWORDS)
    date_columns = _find_columns(headers, DATE_KEYWORDS)

    ticket_index = ticket_columns[0] if ticket_columns else None
    subject_index = subject_columns[0] if subject_columns else None
    requester_index = requester_columns[0] if requester_columns else None
    status_index = status_columns[0] if status_columns else None
    priority_index = priority_columns[0] if priority_columns else None
    date_index = date_columns[0] if date_columns else None

    total_tickets = 0
    open_tickets = 0
    closed_tickets = 0
    high_priority = 0
    missing_ticket_id = 0
    missing_subject = 0
    missing_requester = 0
    missing_status = 0
    old_open_tickets = 0

    today = datetime.now().date()
    requester_counts = {}
    warnings = []

    for row_number, row in enumerate(rows, start=2):
        if not any(_normalize(cell) for cell in row):
            continue

        total_tickets += 1

        ticket_id = _normalize(row[ticket_index]) if ticket_index is not None and ticket_index < len(row) else ""
        subject = _normalize(row[subject_index]) if subject_index is not None and subject_index < len(row) else ""
        requester = _normalize(row[requester_index]) if requester_index is not None and requester_index < len(row) else ""
        status = _normalize(row[status_index]) if status_index is not None and status_index < len(row) else ""
        priority = _normalize(row[priority_index]) if priority_index is not None and priority_index < len(row) else ""
        created_date = _parse_date(row[date_index]) if date_index is not None and date_index < len(row) else None

        if not ticket_id:
            missing_ticket_id += 1
            warnings.append(f"- Row {row_number}: Missing ticket ID/reference.")

        if not subject:
            missing_subject += 1
            warnings.append(f"- Row {row_number}: Missing ticket subject/description.")

        if requester:
            requester_counts[requester] = requester_counts.get(requester, 0) + 1
        else:
            missing_requester += 1
            warnings.append(f"- Row {row_number}: Missing requester/user.")

        if not status:
            missing_status += 1
            warnings.append(f"- Row {row_number}: Missing status.")

        if _is_closed(status):
            closed_tickets += 1
        elif _is_open(status) or status:
            open_tickets += 1

        if _is_high_priority(priority):
            high_priority += 1
            warnings.append(f"- Row {row_number}: High priority ticket detected.")

        if created_date and not _is_closed(status):
            age_days = (today - created_date).days
            if age_days > 7:
                old_open_tickets += 1
                warnings.append(f"- Row {row_number}: Open ticket older than 7 days.")

    return {
        "ticket_columns": [headers[i] for i in ticket_columns if i < len(headers)],
        "subject_columns": [headers[i] for i in subject_columns if i < len(headers)],
        "requester_columns": [headers[i] for i in requester_columns if i < len(headers)],
        "status_columns": [headers[i] for i in status_columns if i < len(headers)],
        "priority_columns": [headers[i] for i in priority_columns if i < len(headers)],
        "date_columns": [headers[i] for i in date_columns if i < len(headers)],
        "total_tickets": total_tickets,
        "open_tickets": open_tickets,
        "closed_tickets": closed_tickets,
        "high_priority": high_priority,
        "missing_ticket_id": missing_ticket_id,
        "missing_subject": missing_subject,
        "missing_requester": missing_requester,
        "missing_status": missing_status,
        "old_open_tickets": old_open_tickets,
        "requester_counts": requester_counts,
        "warnings": warnings[:80],
    }


def _format_summary(file_type, summary):
    lines = [
        f"- Type: {file_type}",
        f"- Total tickets detected: {summary['total_tickets']}",
        f"- Open tickets: {summary['open_tickets']}",
        f"- Closed/resolved tickets: {summary['closed_tickets']}",
        f"- High priority tickets: {summary['high_priority']}",
        f"- Open tickets older than 7 days: {summary['old_open_tickets']}",
        f"- Missing ticket ID rows: {summary['missing_ticket_id']}",
        f"- Missing subject rows: {summary['missing_subject']}",
        f"- Missing requester rows: {summary['missing_requester']}",
        f"- Missing status rows: {summary['missing_status']}",
        f"- Ticket ID columns: {', '.join(map(str, summary['ticket_columns'])) if summary['ticket_columns'] else 'NONE'}",
        f"- Subject columns: {', '.join(map(str, summary['subject_columns'])) if summary['subject_columns'] else 'NONE'}",
        f"- Requester columns: {', '.join(map(str, summary['requester_columns'])) if summary['requester_columns'] else 'NONE'}",
        f"- Status columns: {', '.join(map(str, summary['status_columns'])) if summary['status_columns'] else 'NONE'}",
        f"- Priority columns: {', '.join(map(str, summary['priority_columns'])) if summary['priority_columns'] else 'NONE'}",
        f"- Date columns: {', '.join(map(str, summary['date_columns'])) if summary['date_columns'] else 'NONE'}",
    ]

    if summary["requester_counts"]:
        lines.append("")
        lines.append("Tickets by requester:")
        for name, count in sorted(summary["requester_counts"].items(), key=lambda item: item[0].lower())[:30]:
            lines.append(f"- {name}: {count}")

    if summary["warnings"]:
        lines.append("")
        lines.append("Helpdesk warnings:")
        lines.extend(summary["warnings"])
    else:
        lines.append("- No obvious helpdesk issues detected.")

    return lines


def _analyze_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]
    summary = _analyze_helpdesk_rows(headers, data_rows)

    return _format_summary("CSV", summary)


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        "- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        summary = _analyze_helpdesk_rows(headers, data_rows)

        for item in _format_summary("XLSX Sheet", summary):
            lines.append(f"  {item}")

    workbook.close()
    return lines


def internal_helpdesk_system():
    project, error = _project()
    if error:
        return error

    files = _find_helpdesk_files(project)

    lines = [
        "INTERNAL HELPDESK SYSTEM — PHASE 350",
        f"Project: {project}",
        "",
        "Mode: read-only helpdesk ticket inspection.",
        "",
    ]

    if not files:
        lines.append("No helpdesk/ticket spreadsheet files found.")
        lines.append("")
        lines.append("File names should include words like:")
        lines.append("- helpdesk")
        lines.append("- support")
        lines.append("- ticket")
        lines.append("- issue")
        lines.append("- request")
        return "\n".join(lines)

    lines.append(f"Helpdesk files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Important:")
    lines.append("- This is a read-only internal helpdesk assistant.")
    lines.append("- It does not create, assign, close, or delete tickets.")
    lines.append("- Ticket actions should be added later through confirm-based workflows.")

    return "\n".join(lines)
