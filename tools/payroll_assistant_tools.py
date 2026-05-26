from pathlib import Path
import csv
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path

SUPPORTED_EXTENSIONS = [".csv", ".xlsx"]
MAX_FILES = 25
MAX_ROWS_TO_SCAN = 1000

PAYROLL_KEYWORDS = [
    "payroll",
    "salary",
    "wage",
    "employee",
    "staff",
    "allowance",
    "deduction",
]

GROSS_KEYWORDS = [
    "gross",
    "basic",
    "salary",
    "wage",
    "earnings",
]

DEDUCTION_KEYWORDS = [
    "deduction",
    "epf",
    "etf",
    "tax",
    "loan",
    "advance",
    "cut",
]

NET_KEYWORDS = [
    "net",
    "payable",
    "take home",
    "takehome",
]

EMPLOYEE_KEYWORDS = [
    "employee",
    "staff",
    "name",
    "worker",
]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/reports",
        "storage/presentations",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_payroll_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            name = file.name.lower()
            if any(keyword in name for keyword in PAYROLL_KEYWORDS):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _normalize(value):
    return str(value or "").strip()


def _normalize_lower(value):
    return _normalize(value).lower()


def _to_number(value):
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    text = (
        text.replace(",", "")
        .replace("Rs.", "")
        .replace("LKR", "")
        .replace("$", "")
        .strip()
    )

    try:
        return float(text)
    except ValueError:
        return None


def _find_columns(headers, keywords):
    indexes = []

    for index, header in enumerate(headers):
        header_text = _normalize_lower(header)
        if any(keyword in header_text for keyword in keywords):
            indexes.append(index)

    return indexes


def _sum_columns(row, indexes):
    total = 0.0
    found = False

    for index in indexes:
        if index >= len(row):
            continue

        number = _to_number(row[index])
        if number is None:
            continue

        total += abs(number)
        found = True

    return total if found else None


def _analyze_payroll_rows(headers, rows):
    employee_columns = _find_columns(headers, EMPLOYEE_KEYWORDS)
    gross_columns = _find_columns(headers, GROSS_KEYWORDS)
    deduction_columns = _find_columns(headers, DEDUCTION_KEYWORDS)
    net_columns = _find_columns(headers, NET_KEYWORDS)

    employee_count = 0
    gross_total = 0.0
    deduction_total = 0.0
    net_total = 0.0
    missing_net = 0
    warnings = []

    for row_number, row in enumerate(rows, start=2):
        if not any(_normalize(cell) for cell in row):
            continue

        employee_count += 1

        gross = _sum_columns(row, gross_columns)
        deductions = _sum_columns(row, deduction_columns)
        net = _sum_columns(row, net_columns)

        if gross is not None:
            gross_total += gross

        if deductions is not None:
            deduction_total += deductions

        if net is not None:
            net_total += net
        else:
            missing_net += 1

        if gross is not None and deductions is not None and net is not None:
            expected_net = gross - deductions
            if abs(expected_net - net) > 1:
                warnings.append(
                    f"- Row {row_number}: Net pay mismatch. Expected {expected_net:,.2f}, found {net:,.2f}."
                )

        if gross is None and net is None:
            warnings.append(
                f"- Row {row_number}: No gross salary or net pay detected."
            )

    return {
        "employee_columns": [headers[i] for i in employee_columns if i < len(headers)],
        "gross_columns": [headers[i] for i in gross_columns if i < len(headers)],
        "deduction_columns": [headers[i] for i in deduction_columns if i < len(headers)],
        "net_columns": [headers[i] for i in net_columns if i < len(headers)],
        "employee_count": employee_count,
        "gross_total": gross_total,
        "deduction_total": deduction_total,
        "net_total": net_total,
        "missing_net": missing_net,
        "warnings": warnings[:50],
    }


def _format_summary(file_type, headers, summary):
    lines = [
        f"- Type: {file_type}",
        f"- Employees/rows detected: {summary['employee_count']}",
        f"- Employee columns: {', '.join(map(str, summary['employee_columns'])) if summary['employee_columns'] else 'NONE'}",
        f"- Gross salary columns: {', '.join(map(str, summary['gross_columns'])) if summary['gross_columns'] else 'NONE'}",
        f"- Deduction columns: {', '.join(map(str, summary['deduction_columns'])) if summary['deduction_columns'] else 'NONE'}",
        f"- Net pay columns: {', '.join(map(str, summary['net_columns'])) if summary['net_columns'] else 'NONE'}",
        f"- Estimated gross payroll: {summary['gross_total']:,.2f}",
        f"- Estimated deductions: {summary['deduction_total']:,.2f}",
        f"- Estimated net payroll: {summary['net_total']:,.2f}",
        f"- Rows missing net pay: {summary['missing_net']}",
    ]

    if not summary["gross_columns"]:
        lines.append("- Warning: No gross/basic salary column detected.")

    if not summary["net_columns"]:
        lines.append("- Warning: No net/payable salary column detected.")

    if summary["warnings"]:
        lines.append("")
        lines.append("Payroll warnings:")
        lines.extend(summary["warnings"])
    else:
        lines.append("- No obvious payroll calculation mismatches detected.")

    return lines


def _analyze_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]
    summary = _analyze_payroll_rows(headers, data_rows)

    return _format_summary("CSV", headers, summary)


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        "- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        summary = _analyze_payroll_rows(headers, data_rows)

        for item in _format_summary("XLSX Sheet", headers, summary):
            lines.append(f"  {item}")

    workbook.close()
    return lines


def payroll_assistant():
    project, error = _project()
    if error:
        return error

    files = _find_payroll_files(project)

    lines = [
        "PAYROLL ASSISTANT — PHASE 346",
        f"Project: {project}",
        "",
        "Mode: read-only payroll inspection.",
        "",
    ]

    if not files:
        lines.append("No payroll spreadsheet files found.")
        lines.append("")
        lines.append("File names should include words like:")
        lines.append("- payroll")
        lines.append("- salary")
        lines.append("- wage")
        lines.append("- employee")
        lines.append("- staff")
        return "\n".join(lines)

    lines.append(f"Payroll files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Important:")
    lines.append("- This is a payroll assistant, not certified payroll/legal advice.")
    lines.append("- It estimates payroll totals from spreadsheet column names.")
    lines.append("- No files were modified.")

    return "\n".join(lines)
