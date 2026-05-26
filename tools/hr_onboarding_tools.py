from pathlib import Path
import csv
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path

SUPPORTED_EXTENSIONS = [".csv", ".xlsx"]
MAX_FILES = 25
MAX_ROWS_TO_SCAN = 1000

HR_KEYWORDS = [
    "hr",
    "human",
    "employee",
    "staff",
    "onboarding",
    "joining",
    "recruit",
]

REQUIRED_FIELDS = [
    "name",
    "email",
    "phone",
    "address",
    "position",
    "department",
    "start date",
    "joining date",
    "salary",
    "id",
    "nic",
    "contract",
    "emergency",
]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/reports",
        "storage/presentations",
    }
    return any(part in skip_dirs for part in path.parts)


def _find_hr_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            name = file.name.lower()
            if any(keyword in name for keyword in HR_KEYWORDS):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _normalize(value):
    return str(value or "").strip()


def _normalize_lower(value):
    return _normalize(value).lower()


def _header_map(headers):
    return {
        index: _normalize_lower(header)
        for index, header in enumerate(headers)
    }


def _field_present(headers, keyword):
    for header in headers:
        if keyword in _normalize_lower(header):
            return True
    return False


def _missing_required_fields(headers):
    return [
        field for field in REQUIRED_FIELDS
        if not _field_present(headers, field)
    ]


def _detect_onboarding_status(headers, rows):
    headers_map = _header_map(headers)
    missing_fields = _missing_required_fields(headers)

    total_candidates = 0
    incomplete_rows = []
    likely_ready = 0

    for row_number, row in enumerate(rows, start=2):
        if not any(_normalize(cell) for cell in row):
            continue

        total_candidates += 1
        row_missing = []

        for index, header in headers_map.items():
            header_text = header
            value = _normalize(row[index] if index < len(row) else "")

            if any(word in header_text for word in ["name", "email", "phone", "position", "department", "start", "joining", "contract"]):
                if not value:
                    row_missing.append(header_text)

        if row_missing:
            incomplete_rows.append(
                f"- Row {row_number}: Missing {', '.join(row_missing[:8])}"
            )
        else:
            likely_ready += 1

    return {
        "total_candidates": total_candidates,
        "likely_ready": likely_ready,
        "incomplete_count": len(incomplete_rows),
        "missing_fields": missing_fields,
        "incomplete_rows": incomplete_rows[:50],
    }


def _format_summary(file_type, headers, summary):
    lines = [
        f"- Type: {file_type}",
        f"- Candidate/employee rows detected: {summary['total_candidates']}",
        f"- Likely ready for onboarding: {summary['likely_ready']}",
        f"- Rows with missing onboarding data: {summary['incomplete_count']}",
    ]

    if summary["missing_fields"]:
        lines.append("")
        lines.append("Missing recommended onboarding columns:")
        for field in summary["missing_fields"]:
            lines.append(f"- {field}")
    else:
        lines.append("- Recommended onboarding columns appear to be present.")

    if summary["incomplete_rows"]:
        lines.append("")
        lines.append("Incomplete onboarding rows:")
        lines.extend(summary["incomplete_rows"])
    else:
        lines.append("- No obvious incomplete onboarding rows detected.")

    return lines


def _analyze_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception as e:
        return [f"- Error reading CSV: {e}"]

    if not rows:
        return ["- Empty CSV file."]

    headers = rows[0]
    data_rows = rows[1:]
    summary = _detect_onboarding_status(headers, data_rows)

    return _format_summary("CSV", headers, summary)


def _analyze_xlsx(file: Path):
    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception as e:
        return [f"- Error reading workbook: {e}"]

    lines = [
        "- Type: XLSX",
        f"- Sheets: {len(workbook.sheetnames)}",
    ]

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]

        rows = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        lines.append("")
        lines.append(f"  Sheet: {sheet_name}")

        if not rows:
            lines.append("  - Empty sheet.")
            continue

        headers = rows[0]
        data_rows = rows[1:]
        summary = _detect_onboarding_status(headers, data_rows)

        for item in _format_summary("XLSX Sheet", headers, summary):
            lines.append(f"  {item}")

    workbook.close()
    return lines


def hr_onboarding_workflow():
    project, error = _project()
    if error:
        return error

    files = _find_hr_files(project)

    lines = [
        "HR ONBOARDING WORKFLOW — PHASE 347",
        f"Project: {project}",
        "",
        "Mode: read-only HR onboarding inspection.",
        "",
    ]

    if not files:
        lines.append("No HR/onboarding spreadsheet files found.")
        lines.append("")
        lines.append("File names should include words like:")
        lines.append("- hr")
        lines.append("- employee")
        lines.append("- staff")
        lines.append("- onboarding")
        lines.append("- joining")
        return "\n".join(lines)

    lines.append(f"HR/onboarding files found: {len(files)}")
    lines.append("")

    for file in files:
        lines.append("=" * 80)
        lines.append(f"FILE: {file.relative_to(project)}")
        lines.append("=" * 80)

        if file.suffix.lower() == ".csv":
            lines.extend(_analyze_csv(file))
        elif file.suffix.lower() == ".xlsx":
            lines.extend(_analyze_xlsx(file))
        else:
            lines.append("- Unsupported file type.")

        lines.append("")

    lines.append("Recommended onboarding checklist:")
    lines.append("- Confirm employee identity details.")
    lines.append("- Confirm role, department, start date, and salary.")
    lines.append("- Collect contract and policy acknowledgements.")
    lines.append("- Prepare system accounts and access permissions.")
    lines.append("- Schedule orientation and first-week review.")
    lines.append("")
    lines.append("Important:")
    lines.append("- This is a read-only HR assistant.")
    lines.append("- It does not create employees or modify records.")
    lines.append("- Final HR decisions must be reviewed by management.")

    return "\n".join(lines)
