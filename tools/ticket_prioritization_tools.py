from pathlib import Path
import csv
from datetime import datetime
from openpyxl import load_workbook

from tools.project_context_tools import get_current_project_path

SUPPORTED_EXTENSIONS = [".csv", ".xlsx"]
MAX_FILES = 25
MAX_ROWS_TO_SCAN = 2000

TICKET_FILE_KEYWORDS = [
    "helpdesk",
    "support",
    "ticket",
    "tickets",
    "issue",
    "issues",
    "request",
    "requests",
]

TICKET_ID_KEYWORDS = ["ticket", "id", "reference", "case"]
SUBJECT_KEYWORDS = ["subject", "title", "issue", "problem", "request", "description"]
REQUESTER_KEYWORDS = ["requester", "employee", "staff", "user", "name", "customer"]
STATUS_KEYWORDS = ["status", "state", "progress"]
PRIORITY_KEYWORDS = ["priority", "severity", "impact", "level"]
DATE_KEYWORDS = ["date", "created", "opened", "reported"]
CATEGORY_KEYWORDS = ["category", "type", "department", "module"]


def _project():
    project = get_current_project_path()
    if not project:
        return None, "No current project selected. Use: use project <name-or-path>"
    return Path(project), None


def _skip(path: Path):
    skip_dirs = {
        ".git",
        "venv",
        "__pycache__",
        "node_modules",
        "vendor",
        "storage/reports",
        "storage/presentations",
    }
    return any(part in skip_dirs for part in path.parts)


def _normalize(value):
    return str(value or "").strip()


def _lower(value):
    return _normalize(value).lower()


def _find_columns(headers, keywords):
    matches = []
    for index, header in enumerate(headers):
        header_text = _lower(header)
        if any(keyword in header_text for keyword in keywords):
            matches.append(index)
    return matches


def _first_index(headers, keywords):
    matches = _find_columns(headers, keywords)
    return matches[0] if matches else None


def _parse_date(value):
    if value is None:
        return None

    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            return None

    text = _normalize(value)
    if not text:
        return None

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d.%m.%Y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _is_closed(status):
    status = _lower(status)
    return any(word in status for word in ["closed", "resolved", "done", "completed", "fixed"])


def _priority_score(priority):
    priority = _lower(priority)

    if any(word in priority for word in ["p0", "critical", "blocker"]):
        return 50
    if any(word in priority for word in ["p1", "urgent", "high", "severe"]):
        return 40
    if any(word in priority for word in ["p2", "medium", "normal"]):
        return 25
    if any(word in priority for word in ["p3", "low", "minor"]):
        return 10

    return 15


def _subject_score(subject):
    subject = _lower(subject)
    score = 0

    critical_terms = [
        "down",
        "crash",
        "failed",
        "failure",
        "cannot login",
        "login issue",
        "payment",
        "billing",
        "security",
        "data loss",
        "not working",
        "production",
        "urgent",
        "blocked",
    ]

    for term in critical_terms:
        if term in subject:
            score += 8

    return min(score, 35)


def _age_score(created_date, status):
    if not created_date or _is_closed(status):
        return 0

    age_days = (datetime.now().date() - created_date).days

    if age_days >= 30:
        return 30
    if age_days >= 14:
        return 22
    if age_days >= 7:
        return 15
    if age_days >= 3:
        return 8

    return 0


def _status_score(status):
    status = _lower(status)

    if _is_closed(status):
        return -100
    if any(word in status for word in ["waiting", "pending", "hold"]):
        return 12
    if any(word in status for word in ["new", "open"]):
        return 10
    if "in progress" in status:
        return 5

    return 6


def _score_ticket(ticket):
    score = 0
    score += _priority_score(ticket["priority"])
    score += _subject_score(ticket["subject"])
    score += _age_score(ticket["created_date"], ticket["status"])
    score += _status_score(ticket["status"])

    if not ticket["requester"]:
        score -= 5

    if score >= 80:
        level = "P0 CRITICAL"
    elif score >= 60:
        level = "P1 HIGH"
    elif score >= 35:
        level = "P2 MEDIUM"
    else:
        level = "P3 LOW"

    return score, level


def _find_ticket_files(project: Path):
    files = []

    for ext in SUPPORTED_EXTENSIONS:
        for file in project.rglob(f"*{ext}"):
            if _skip(file) or not file.is_file():
                continue

            name = file.name.lower()
            if any(keyword in name for keyword in TICKET_FILE_KEYWORDS):
                files.append(file)

    return sorted(files)[:MAX_FILES]


def _rows_to_tickets(headers, rows, source_name):
    ticket_index = _first_index(headers, TICKET_ID_KEYWORDS)
    subject_index = _first_index(headers, SUBJECT_KEYWORDS)
    requester_index = _first_index(headers, REQUESTER_KEYWORDS)
    status_index = _first_index(headers, STATUS_KEYWORDS)
    priority_index = _first_index(headers, PRIORITY_KEYWORDS)
    date_index = _first_index(headers, DATE_KEYWORDS)
    category_index = _first_index(headers, CATEGORY_KEYWORDS)

    tickets = []

    for row_number, row in enumerate(rows, start=2):
        if not any(_normalize(cell) for cell in row):
            continue

        ticket = {
            "source": source_name,
            "row": row_number,
            "ticket_id": _normalize(row[ticket_index]) if ticket_index is not None and ticket_index < len(row) else "",
            "subject": _normalize(row[subject_index]) if subject_index is not None and subject_index < len(row) else "",
            "requester": _normalize(row[requester_index]) if requester_index is not None and requester_index < len(row) else "",
            "status": _normalize(row[status_index]) if status_index is not None and status_index < len(row) else "",
            "priority": _normalize(row[priority_index]) if priority_index is not None and priority_index < len(row) else "",
            "category": _normalize(row[category_index]) if category_index is not None and category_index < len(row) else "",
            "created_date": _parse_date(row[date_index]) if date_index is not None and date_index < len(row) else None,
        }

        if _is_closed(ticket["status"]):
            continue

        ticket["score"], ticket["level"] = _score_ticket(ticket)
        tickets.append(ticket)

    return tickets


def _read_csv(file: Path):
    rows = []

    try:
        with file.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                if index >= MAX_ROWS_TO_SCAN:
                    break
                rows.append(row)
    except Exception:
        return []

    if not rows:
        return []

    return _rows_to_tickets(rows[0], rows[1:], file.name)


def _read_xlsx(file: Path):
    tickets = []

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        return tickets

    for sheet_name in workbook.sheetnames[:5]:
        sheet = workbook[sheet_name]
        rows = []

        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS_TO_SCAN:
                break
            rows.append(row)

        if rows:
            source_name = f"{file.name} / {sheet_name}"
            tickets.extend(_rows_to_tickets(rows[0], rows[1:], source_name))

    workbook.close()
    return tickets


def ticket_prioritization_engine():
    project, error = _project()
    if error:
        return error

    files = _find_ticket_files(project)

    lines = [
        "TICKET PRIORITIZATION ENGINE — PHASE 351",
        f"Project: {project}",
        "",
        "Mode: read-only ticket scoring and prioritization.",
        "",
    ]

    if not files:
        lines.append("No ticket/helpdesk spreadsheet files found.")
        lines.append("")
        lines.append("Supported files: .csv, .xlsx")
        lines.append("File names should include: ticket, helpdesk, support, issue, or request.")
        return "\n".join(lines)

    tickets = []

    for file in files:
        if file.suffix.lower() == ".csv":
            tickets.extend(_read_csv(file))
        elif file.suffix.lower() == ".xlsx":
            tickets.extend(_read_xlsx(file))

    if not tickets:
        lines.append("No open tickets detected for prioritization.")
        return "\n".join(lines)

    tickets = sorted(tickets, key=lambda item: item["score"], reverse=True)

    lines.append(f"Ticket files scanned: {len(files)}")
    lines.append(f"Open tickets prioritized: {len(tickets)}")
    lines.append("")
    lines.append("TOP PRIORITY TICKETS")
    lines.append("")

    for index, ticket in enumerate(tickets[:25], start=1):
        created = ticket["created_date"].isoformat() if ticket["created_date"] else "Unknown"

        lines.append(f"{index}. {ticket['level']} | Score: {ticket['score']}")
        lines.append(f"   Ticket: {ticket['ticket_id'] or 'N/A'}")
        lines.append(f"   Subject: {ticket['subject'] or 'N/A'}")
        lines.append(f"   Requester: {ticket['requester'] or 'N/A'}")
        lines.append(f"   Status: {ticket['status'] or 'N/A'}")
        lines.append(f"   Priority: {ticket['priority'] or 'N/A'}")
        lines.append(f"   Category: {ticket['category'] or 'N/A'}")
        lines.append(f"   Created: {created}")
        lines.append(f"   Source: {ticket['source']} row {ticket['row']}")
        lines.append("")

    lines.append("Scoring model:")
    lines.append("- Priority/severity increases score.")
    lines.append("- Old unresolved tickets increase score.")
    lines.append("- Critical subject terms increase score.")
    lines.append("- Closed/resolved tickets are ignored.")
    lines.append("")
    lines.append("Safety:")
    lines.append("- Read-only analysis only.")
    lines.append("- No ticket assignment, closing, editing, or deletion is performed.")

    return "\n".join(lines)
